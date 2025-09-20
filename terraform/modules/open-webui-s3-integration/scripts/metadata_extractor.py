#!/usr/bin/env python3
"""
Metadata Extractor for Open WebUI Document Storage

This module extracts metadata from various file types including
text content, document properties, EXIF data, and generates embeddings.
"""

import os
import json
import logging
import mimetypes
from datetime import datetime
from typing import Dict, List, Optional, Any, Tuple
from pathlib import Path
import hashlib

# Document processing libraries
try:
    import PyPDF2
    import pdfplumber
    HAS_PDF = True
except ImportError:
    HAS_PDF = False

try:
    from docx import Document as DocxDocument
    HAS_DOCX = True
except ImportError:
    HAS_DOCX = False

try:
    import openpyxl
    HAS_EXCEL = True
except ImportError:
    HAS_EXCEL = False

try:
    from PIL import Image
    from PIL.ExifTags import TAGS
    HAS_PIL = True
except ImportError:
    HAS_PIL = False

try:
    import eyed3
    HAS_AUDIO = True
except ImportError:
    HAS_AUDIO = False

try:
    import pytesseract
    HAS_OCR = True
except ImportError:
    HAS_OCR = False

logger = logging.getLogger(__name__)

class MetadataExtractor:
    def __init__(self, config_path: str = "/app/config/s3_config.json"):
        """Initialize metadata extractor with configuration"""
        self.config = self._load_config(config_path)
        self.extract_text = self.config['indexing']['extract_text']
        self.generate_thumbnails = self.config['indexing']['generate_thumbnails']
        self.extract_metadata = self.config['indexing']['extract_metadata']
        self.ocr_enabled = self.config['indexing']['ocr_enabled']
        
    def _load_config(self, config_path: str) -> Dict:
        """Load configuration from file"""
        try:
            with open(config_path, 'r') as f:
                return json.load(f)
        except FileNotFoundError:
            logger.error(f"Configuration file not found: {config_path}")
            raise
        except json.JSONDecodeError as e:
            logger.error(f"Invalid JSON in configuration file: {e}")
            raise
    
    def calculate_file_hash(self, file_path: str) -> str:
        """Calculate SHA-256 hash of file"""
        sha256_hash = hashlib.sha256()
        
        with open(file_path, "rb") as f:
            for chunk in iter(lambda: f.read(4096), b""):
                sha256_hash.update(chunk)
        
        return sha256_hash.hexdigest()
    
    def get_basic_file_info(self, file_path: str) -> Dict:
        """Extract basic file information"""
        try:
            file_stat = os.stat(file_path)
            file_path_obj = Path(file_path)
            
            # Detect MIME type
            mime_type, encoding = mimetypes.guess_type(file_path)
            
            return {
                'filename': file_path_obj.name,
                'file_extension': file_path_obj.suffix.lower(),
                'file_size': file_stat.st_size,
                'mime_type': mime_type,
                'encoding': encoding,
                'created_time': datetime.fromtimestamp(file_stat.st_ctime).isoformat(),
                'modified_time': datetime.fromtimestamp(file_stat.st_mtime).isoformat(),
                'file_hash': self.calculate_file_hash(file_path)
            }
            
        except Exception as e:
            logger.error(f"Error extracting basic file info: {e}")
            return {}
    
    def extract_pdf_metadata(self, file_path: str) -> Dict:
        """Extract metadata from PDF files"""
        metadata = {'text_content': '', 'document_info': {}, 'page_count': 0}
        
        if not HAS_PDF:
            logger.warning("PDF processing libraries not available")
            return metadata
        
        try:
            # Extract document info and text with pdfplumber (more reliable)
            with pdfplumber.open(file_path) as pdf:
                metadata['page_count'] = len(pdf.pages)
                metadata['document_info'] = pdf.metadata or {}
                
                if self.extract_text:
                    text_content = []
                    for page in pdf.pages:
                        page_text = page.extract_text()
                        if page_text:
                            text_content.append(page_text)
                    
                    metadata['text_content'] = '\n'.join(text_content)
            
            # Fallback to PyPDF2 for additional metadata
            try:
                with open(file_path, 'rb') as file:
                    pdf_reader = PyPDF2.PdfReader(file)
                    
                    if pdf_reader.metadata:
                        for key, value in pdf_reader.metadata.items():
                            if key.startswith('/'):
                                clean_key = key[1:]  # Remove leading slash
                                metadata['document_info'][clean_key] = str(value)
                    
                    # Update page count if not set
                    if not metadata['page_count']:
                        metadata['page_count'] = len(pdf_reader.pages)
                        
            except Exception as e:
                logger.debug(f"PyPDF2 fallback failed: {e}")
            
        except Exception as e:
            logger.error(f"Error extracting PDF metadata: {e}")
        
        return metadata
    
    def extract_docx_metadata(self, file_path: str) -> Dict:
        """Extract metadata from DOCX files"""
        metadata = {'text_content': '', 'document_properties': {}}
        
        if not HAS_DOCX:
            logger.warning("DOCX processing library not available")
            return metadata
        
        try:
            doc = DocxDocument(file_path)
            
            # Extract text content
            if self.extract_text:
                paragraphs = []
                for paragraph in doc.paragraphs:
                    if paragraph.text.strip():
                        paragraphs.append(paragraph.text)
                
                metadata['text_content'] = '\n'.join(paragraphs)
            
            # Extract document properties
            if self.extract_metadata:
                core_props = doc.core_properties
                
                properties = {}
                for prop in ['author', 'category', 'comments', 'content_status', 
                           'created', 'identifier', 'keywords', 'language', 
                           'last_modified_by', 'last_printed', 'modified', 
                           'revision', 'subject', 'title', 'version']:
                    try:
                        value = getattr(core_props, prop, None)
                        if value is not None:
                            if hasattr(value, 'isoformat'):
                                properties[prop] = value.isoformat()
                            else:
                                properties[prop] = str(value)
                    except:
                        pass
                
                metadata['document_properties'] = properties
            
        except Exception as e:
            logger.error(f"Error extracting DOCX metadata: {e}")
        
        return metadata
    
    def extract_excel_metadata(self, file_path: str) -> Dict:
        """Extract metadata from Excel files"""
        metadata = {'text_content': '', 'document_properties': {}, 'sheet_info': []}
        
        if not HAS_EXCEL:
            logger.warning("Excel processing library not available")
            return metadata
        
        try:
            workbook = openpyxl.load_workbook(file_path, data_only=True)
            
            # Extract document properties
            if self.extract_metadata:
                props = workbook.properties
                
                properties = {}
                for prop in ['creator', 'title', 'description', 'subject', 
                           'identifier', 'language', 'created', 'modified', 
                           'lastModifiedBy', 'category', 'contentStatus', 
                           'version', 'revision', 'keywords']:
                    try:
                        value = getattr(props, prop, None)
                        if value is not None:
                            if hasattr(value, 'isoformat'):
                                properties[prop] = value.isoformat()
                            else:
                                properties[prop] = str(value)
                    except:
                        pass
                
                metadata['document_properties'] = properties
            
            # Extract sheet information and text content
            sheet_info = []
            text_content = []
            
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                
                # Count rows and columns with data
                max_row = sheet.max_row
                max_col = sheet.max_column
                
                sheet_data = {
                    'name': sheet_name,
                    'max_row': max_row,
                    'max_column': max_col
                }
                
                # Extract text content if enabled
                if self.extract_text and max_row <= 1000:  # Limit for performance
                    sheet_text = []
                    for row in sheet.iter_rows(max_row=min(max_row, 100), values_only=True):
                        row_text = []
                        for cell in row:
                            if cell is not None:
                                row_text.append(str(cell))
                        if row_text:
                            sheet_text.append(' | '.join(row_text))
                    
                    if sheet_text:
                        text_content.append(f"Sheet: {sheet_name}\n" + '\n'.join(sheet_text))
                
                sheet_info.append(sheet_data)
            
            metadata['sheet_info'] = sheet_info
            if text_content:
                metadata['text_content'] = '\n\n'.join(text_content)
            
        except Exception as e:
            logger.error(f"Error extracting Excel metadata: {e}")
        
        return metadata
    
    def extract_image_metadata(self, file_path: str) -> Dict:
        """Extract metadata from image files"""
        metadata = {'image_info': {}, 'exif_data': {}}
        
        if not HAS_PIL:
            logger.warning("PIL library not available for image processing")
            return metadata
        
        try:
            with Image.open(file_path) as image:
                # Basic image information
                metadata['image_info'] = {
                    'format': image.format,
                    'mode': image.mode,
                    'size': image.size,
                    'width': image.width,
                    'height': image.height
                }
                
                # Extract EXIF data if available
                if self.extract_metadata and hasattr(image, '_getexif'):
                    exif_data = image._getexif()
                    
                    if exif_data:
                        exif_dict = {}
                        for tag_id, value in exif_data.items():
                            tag = TAGS.get(tag_id, tag_id)
                            try:
                                # Convert bytes to string if needed
                                if isinstance(value, bytes):
                                    value = value.decode('utf-8', errors='ignore')
                                exif_dict[tag] = str(value)
                            except:
                                pass
                        
                        metadata['exif_data'] = exif_dict
                
                # Perform OCR if enabled and image is suitable
                if self.ocr_enabled and HAS_OCR:
                    try:
                        ocr_text = pytesseract.image_to_string(image)
                        if ocr_text.strip():
                            metadata['ocr_text'] = ocr_text.strip()
                    except Exception as e:
                        logger.debug(f"OCR failed for {file_path}: {e}")
            
        except Exception as e:
            logger.error(f"Error extracting image metadata: {e}")
        
        return metadata
    
    def extract_audio_metadata(self, file_path: str) -> Dict:
        """Extract metadata from audio files"""
        metadata = {'audio_info': {}, 'tags': {}}
        
        if not HAS_AUDIO:
            logger.warning("Audio processing library not available")
            return metadata
        
        try:
            audiofile = eyed3.load(file_path)
            
            if audiofile and audiofile.info:
                # Basic audio information
                metadata['audio_info'] = {
                    'duration_seconds': audiofile.info.time_secs,
                    'bitrate': audiofile.info.bit_rate[1] if audiofile.info.bit_rate else None,
                    'sample_rate': audiofile.info.sample_freq,
                    'mode': audiofile.info.mode
                }
                
                # Extract ID3 tags
                if audiofile.tag and self.extract_metadata:
                    tags = {}
                    
                    for attr in ['title', 'artist', 'album', 'album_artist', 
                               'composer', 'genre', 'recording_date', 'release_date',
                               'track_num', 'disc_num']:
                        try:
                            value = getattr(audiofile.tag, attr, None)
                            if value is not None:
                                if hasattr(value, 'isoformat'):
                                    tags[attr] = value.isoformat()
                                else:
                                    tags[attr] = str(value)
                        except:
                            pass
                    
                    metadata['tags'] = tags
            
        except Exception as e:
            logger.error(f"Error extracting audio metadata: {e}")
        
        return metadata
    
    def extract_text_file_content(self, file_path: str) -> Dict:
        """Extract content from text files"""
        metadata = {'text_content': '', 'encoding': None, 'line_count': 0}
        
        if not self.extract_text:
            return metadata
        
        try:
            # Try different encodings
            encodings = ['utf-8', 'utf-16', 'latin-1', 'cp1252']
            
            for encoding in encodings:
                try:
                    with open(file_path, 'r', encoding=encoding) as f:
                        content = f.read()
                        
                    metadata['text_content'] = content
                    metadata['encoding'] = encoding
                    metadata['line_count'] = len(content.splitlines())
                    break
                    
                except UnicodeDecodeError:
                    continue
            
            if not metadata['text_content']:
                logger.warning(f"Could not decode text file: {file_path}")
            
        except Exception as e:
            logger.error(f"Error extracting text file content: {e}")
        
        return metadata
    
    def generate_thumbnail(self, file_path: str, output_path: str, size: Tuple[int, int] = (200, 200)) -> Dict:
        """Generate thumbnail for supported file types"""
        if not self.generate_thumbnails or not HAS_PIL:
            return {'thumbnail_generated': False, 'reason': 'Thumbnail generation disabled or PIL not available'}
        
        try:
            file_ext = Path(file_path).suffix.lower()
            
            # Handle image files
            if file_ext in ['.jpg', '.jpeg', '.png', '.gif', '.bmp', '.tiff', '.webp']:
                with Image.open(file_path) as image:
                    # Convert to RGB if necessary
                    if image.mode in ('RGBA', 'LA', 'P'):
                        image = image.convert('RGB')
                    
                    # Generate thumbnail
                    image.thumbnail(size, Image.Resampling.LANCZOS)
                    image.save(output_path, 'JPEG', quality=85)
                    
                    return {
                        'thumbnail_generated': True,
                        'thumbnail_path': output_path,
                        'thumbnail_size': image.size
                    }
            
            # Handle PDF files (first page)
            elif file_ext == '.pdf' and HAS_PDF:
                try:
                    import fitz  # PyMuPDF
                    
                    doc = fitz.open(file_path)
                    page = doc[0]  # First page
                    
                    # Render page as image
                    mat = fitz.Matrix(1.0, 1.0)  # Scale factor
                    pix = page.get_pixmap(matrix=mat)
                    
                    # Convert to PIL Image
                    img_data = pix.tobytes("ppm")
                    image = Image.open(io.BytesIO(img_data))
                    
                    # Generate thumbnail
                    image.thumbnail(size, Image.Resampling.LANCZOS)
                    image.save(output_path, 'JPEG', quality=85)
                    
                    doc.close()
                    
                    return {
                        'thumbnail_generated': True,
                        'thumbnail_path': output_path,
                        'thumbnail_size': image.size
                    }
                    
                except ImportError:
                    logger.debug("PyMuPDF not available for PDF thumbnail generation")
            
            return {'thumbnail_generated': False, 'reason': f'Unsupported file type for thumbnails: {file_ext}'}
            
        except Exception as e:
            logger.error(f"Error generating thumbnail: {e}")
            return {'thumbnail_generated': False, 'error': str(e)}
    
    def extract_metadata(self, file_path: str) -> Dict:
        """Extract comprehensive metadata from file"""
        logger.info(f"Extracting metadata from: {file_path}")
        
        if not os.path.exists(file_path):
            return {
                'success': False,
                'error': 'File not found'
            }
        
        try:
            # Start with basic file information
            metadata = {
                'extraction_time': datetime.utcnow().isoformat(),
                'file_info': self.get_basic_file_info(file_path),
                'content_metadata': {}
            }
            
            # Determine file type and extract specific metadata
            file_ext = Path(file_path).suffix.lower()
            mime_type = metadata['file_info'].get('mime_type', '')
            
            # PDF files
            if file_ext == '.pdf' or 'pdf' in mime_type:
                metadata['content_metadata'].update(self.extract_pdf_metadata(file_path))
            
            # Microsoft Word documents
            elif file_ext in ['.docx', '.doc'] or 'wordprocessingml' in mime_type:
                metadata['content_metadata'].update(self.extract_docx_metadata(file_path))
            
            # Excel files
            elif file_ext in ['.xlsx', '.xls'] or 'spreadsheetml' in mime_type:
                metadata['content_metadata'].update(self.extract_excel_metadata(file_path))
            
            # Image files
            elif file_ext in ['.jpg', '.jpeg', '.png', '.gif', '.bmp', '.tiff', '.webp'] or 'image/' in mime_type:
                metadata['content_metadata'].update(self.extract_image_metadata(file_path))
            
            # Audio files
            elif file_ext in ['.mp3', '.wav', '.flac', '.ogg'] or 'audio/' in mime_type:
                metadata['content_metadata'].update(self.extract_audio_metadata(file_path))
            
            # Text files
            elif file_ext in ['.txt', '.md', '.csv', '.json', '.xml', '.html', '.htm'] or 'text/' in mime_type:
                metadata['content_metadata'].update(self.extract_text_file_content(file_path))
            
            # Generate thumbnail if requested
            if self.generate_thumbnails:
                thumbnail_path = f"/tmp/thumbnail_{metadata['file_info']['file_hash']}.jpg"
                thumbnail_result = self.generate_thumbnail(file_path, thumbnail_path)
                metadata['thumbnail'] = thumbnail_result
            
            # Calculate content statistics
            text_content = metadata['content_metadata'].get('text_content', '')
            if text_content:
                metadata['content_stats'] = {
                    'character_count': len(text_content),
                    'word_count': len(text_content.split()),
                    'line_count': len(text_content.splitlines()),
                    'has_content': bool(text_content.strip())
                }
            
            logger.info(f"Successfully extracted metadata from: {file_path}")
            
            return {
                'success': True,
                'metadata': metadata
            }
            
        except Exception as e:
            logger.error(f"Error extracting metadata from {file_path}: {e}")
            return {
                'success': False,
                'error': str(e)
            }
    
    def health_check(self) -> Dict:
        """Perform health check on metadata extraction capabilities"""
        try:
            capabilities = {
                'pdf_processing': HAS_PDF,
                'docx_processing': HAS_DOCX,
                'excel_processing': HAS_EXCEL,
                'image_processing': HAS_PIL,
                'audio_processing': HAS_AUDIO,
                'ocr_processing': HAS_OCR and self.ocr_enabled,
                'text_extraction': self.extract_text,
                'thumbnail_generation': self.generate_thumbnails,
                'metadata_extraction': self.extract_metadata
            }
            
            # Count available capabilities
            available_count = sum(1 for v in capabilities.values() if v)
            total_count = len(capabilities)
            
            return {
                'success': True,
                'capabilities': capabilities,
                'available_processors': available_count,
                'total_processors': total_count,
                'coverage_percentage': (available_count / total_count) * 100,
                'timestamp': datetime.utcnow().isoformat()
            }
            
        except Exception as e:
            logger.error(f"Metadata extractor health check failed: {e}")
            return {
                'success': False,
                'error': str(e)
            }